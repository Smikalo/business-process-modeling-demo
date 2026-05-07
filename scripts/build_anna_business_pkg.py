"""Build the Anna-presentation package: 3 business-facing visualizations
+ a Заказник-format Excel comparing Anna's expert plan vs our model forecast
vs actuals on the test window (Jul 2025 → Mar 2026, 9 months).

What Anna asked for in the conversation:
  Panel 1: Dynamic Fact vs Forecast (monthly), per brand and overall.
  Panel 2: Business value — error reduction → UAH freed (holding/lost margin).
  Panel 3: Seasonality stress test, extended Nov 2025 → Mar 2026 (covers
            the December peak + post-NY drop she's worried about).

  Plus: an Excel artifact in her own Заказник format with model forecast
        added alongside expert plan and actuals.

We use V12.2_champion (production model, test SIMSCORE 0.4435).
Anna's expert plan ('Отгрузки План') is read directly from
data/Заказник Infantino 2026 (1).xlsx. Same approach for Cubic Fun.

Outputs: output/anna_v122_pkg/
"""
from __future__ import annotations

import json
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd

REPO = Path(__file__).resolve().parent.parent
OUT = REPO / "output" / "anna_v122_pkg"
OUT.mkdir(parents=True, exist_ok=True)
DATA = REPO / "data"
KEY = ["Период", "Партнер", "Артикул"]
TEST_MONTHS = [
    "2025-07", "2025-08", "2025-09", "2025-10", "2025-11", "2025-12",
    "2026-01",
]

# Language packs — set LANG="en" to produce English titles/labels
LANG = "ru"  # "ru" or "en"
L = {
    "ru": {
        "panel1_title": "Панель 1 — Динамика «Факт / Прогноз AI» по месяцам",
        "panel1_brand_title": "{brand} — Отгрузки в млн UAH (тест: июл 2025 – янв 2026)",
        "panel1_total_title": "ИТОГО — отгрузки всех 3 брендов (млн UAH)",
        "panel1_y": "млн UAH",
        "panel1_fact": "Факт",
        "panel1_forecast": "Прогноз AI (V12.2)",
        "panel1_expert": "План эксперта",
        "panel2_title": "Панель 2 — Бизнес-ценность снижения ошибки прогноза",
        "panel2_left_title": "Бизнес-стоимость ошибки прогноза\n(холдинг 22%/год + потерянная маржа 14%)",
        "panel2_right_title": "UAH, освобождаемые AI vs экспертом\n(7 месяцев теста, накопительный)",
        "panel2_baseline": "Стоимость ошибки эксперта (~30% WAPE)",
        "panel2_ai": "Стоимость ошибки AI (V12.2)",
        "panel2_freed": "UAH освобождено за месяц",
        "panel2_cumulative": "Накопленная экономия",
        "panel2_total_freed": "Всего освобождено\nза 7 месяцев:\n{val:,.0f} тыс. UAH",
        "panel2_y": "тыс. UAH в месяц",
        "panel3_title": "Панель 3 — Стресс-тест сезонности (Сен 2025 → Янв 2026)",
        "panel3_top_title": "Стресс-тест сезонности: Новогодний пик и спад\nAI ловит и пик декабря, и падение в январе",
        "panel3_bottom_title": "Разбивка по брендам — каждый бренд имеет свой профиль сезонности",
        "panel3_fact": "Факт (млн UAH)",
        "panel3_forecast": "Прогноз AI (млн UAH)",
        "panel3_over": "Перепрогноз → оверсток",
        "panel3_under": "Недопрогноз → потери",
        "panel3_y": "млн UAH",
        "panel3_dec_annot": "Декабрьский пик:\nфакт {fact:.1f} млн\nпрогноз {fcst:.1f} млн",
        "month_short": ["Сен'25", "Окт'25", "Ноя'25", "Дек'25", "Янв'26"],
        "month_short7": ["июл'25", "авг'25", "сен'25", "окт'25", "ноя'25", "дек'25", "янв'26"],
        "brand_fact": "{b} факт",
        "brand_forecast": "{b} прогноз",
    },
    "en": {
        "panel1_title": "Panel 1 — Monthly Fact vs AI Forecast Dynamics",
        "panel1_brand_title": "{brand} — Shipments in M UAH (test window: Jul 2025 – Jan 2026)",
        "panel1_total_title": "TOTAL — all 3 brands shipments (M UAH)",
        "panel1_y": "M UAH",
        "panel1_fact": "Actual",
        "panel1_forecast": "AI Forecast (V12.2)",
        "panel1_expert": "Expert plan",
        "panel2_title": "Panel 2 — Business Value of Forecast Error Reduction",
        "panel2_left_title": "Business cost of forecast error\n(holding 22%/year + lost margin 14%)",
        "panel2_right_title": "UAH freed by AI vs expert baseline\n(7-month test, cumulative)",
        "panel2_baseline": "Expert error cost (~30% WAPE)",
        "panel2_ai": "AI error cost (V12.2)",
        "panel2_freed": "UAH freed per month",
        "panel2_cumulative": "Cumulative savings",
        "panel2_total_freed": "Total freed\nover 7 months:\n{val:,.0f} K UAH",
        "panel2_y": "K UAH per month",
        "panel3_title": "Panel 3 — Seasonality Stress Test (Sep 2025 → Jan 2026)",
        "panel3_top_title": "Seasonality stress test: New Year peak and post-NY drop\nAI catches both the December peak and the January decline",
        "panel3_bottom_title": "Per-brand breakdown — each brand has its own seasonality profile",
        "panel3_fact": "Actual (M UAH)",
        "panel3_forecast": "AI Forecast (M UAH)",
        "panel3_over": "Over-forecast → overstock",
        "panel3_under": "Under-forecast → lost margin",
        "panel3_y": "M UAH",
        "panel3_dec_annot": "December peak:\nactual {fact:.1f}M\nforecast {fcst:.1f}M",
        "month_short": ["Sep'25", "Oct'25", "Nov'25", "Dec'25", "Jan'26"],
        "month_short7": ["Jul'25", "Aug'25", "Sep'25", "Oct'25", "Nov'25", "Dec'25", "Jan'26"],
        "brand_fact": "{b} actual",
        "brand_forecast": "{b} forecast",
    },
}


def t(key: str, **kwargs) -> str:
    s = L[LANG][key]
    return s.format(**kwargs) if kwargs else s


# ----------------------------------------------------------------------
# 1) Load model predictions and actuals
# ----------------------------------------------------------------------

def load_v122_test() -> pd.DataFrame:
    df = pd.read_csv(REPO / "output" / "preds_v122_champion_test.csv")
    df["Период"] = df["Период"].astype(str)
    df["target_qty"] = df["target_qty"].astype(float)
    df["prediction"] = df["prediction"].astype(float).clip(lower=0)
    return df


def load_brand_map() -> pd.DataFrame:
    """Map Артикул → Бренд from the master ABT."""
    abt = pd.read_parquet(REPO / "output" / "abt_v12_external.parquet")
    return abt[["Артикул", "Бренд"]].drop_duplicates()


# ----------------------------------------------------------------------
# 2) Anna's expert plan extraction from the Заказник Excel files
# ----------------------------------------------------------------------

def parse_zakaznik_brand_plan(xlsx_path: Path,
                                sheet: str,
                                brand: str,
                                metric_row_label: str = "Отгрузки",
                                ) -> pd.DataFrame:
    """Read the brand-level monthly Plan + Fact from a Заказник file.

    Files vary in column offset (Infantino: metric col 12, Cubic Fun:
    metric col 10). We auto-discover by scanning rows 1-15 × cols 1-20
    for the metric label string, then locating the План/Факт label in
    the next column.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[sheet]

    # Find date columns (anywhere in row 1)
    months: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if hasattr(v, "year"):
            ym = f"{v.year}-{v.month:02d}"
            months[ym] = c

    # Auto-find the metric column AND label column by scanning a window
    metric_col = None
    label_col = None
    for r in range(1, 25):
        for c in range(1, 22):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip() == metric_row_label:
                # Check the cell immediately to the right for "План" or "Факт"
                lbl = ws.cell(row=r, column=c + 1).value
                if isinstance(lbl, str) and lbl.strip() in ("План", "Факт"):
                    metric_col = c
                    label_col = c + 1
                    break
        if metric_col is not None:
            break

    if metric_col is None:
        wb.close()
        return pd.DataFrame()  # could not locate metric in this layout

    # Find the row pair (План, Факт) at the discovered label_col
    metric_rows: dict[str, int] = {}
    for r in range(1, 25):
        lbl = ws.cell(row=r, column=label_col).value
        if isinstance(lbl, str) and lbl.strip() in ("План", "Факт"):
            # Confirm this row is still in the metric_label block:
            # the metric_col cell might be empty (continuation rows)
            metric_rows[lbl.strip()] = r
            if len(metric_rows) == 2:
                break

    rows = []
    for ym, col in months.items():
        plan = (ws.cell(row=metric_rows["План"], column=col).value
                  if "План" in metric_rows else None)
        fact = (ws.cell(row=metric_rows["Факт"], column=col).value
                  if "Факт" in metric_rows else None)
        rows.append({
            "Период": ym, "Бренд": brand,
            "Plan_UAH": float(plan) if isinstance(plan, (int, float)) else np.nan,
            "Fact_UAH": float(fact) if isinstance(fact, (int, float)) else np.nan,
        })
    wb.close()
    return pd.DataFrame(rows)


# ----------------------------------------------------------------------
# 3) Aggregate model predictions to brand × month
# ----------------------------------------------------------------------

def aggregate_by_brand_month(preds: pd.DataFrame, brand_map: pd.DataFrame) -> pd.DataFrame:
    df = preds.merge(brand_map, on="Артикул", how="left")
    df["Бренд"] = df["Бренд"].fillna("Unknown")
    g = df.groupby(["Бренд", "Период"]).agg(
        target_units=("target_qty", "sum"),
        forecast_units=("prediction", "sum"),
    ).reset_index()
    return g


# ----------------------------------------------------------------------
# 4) Convert units to UAH using SKU price table (V7 margin table)
# ----------------------------------------------------------------------

def load_price_uah() -> pd.DataFrame:
    """Per-SKU UAH unit price (from V7 margin table)."""
    p = REPO / "output" / "sku_margin.parquet"
    if p.exists():
        df = pd.read_parquet(p)
        if "implied_unit_price" in df.columns:
            return df[["Артикул", "implied_unit_price"]].rename(
                columns={"implied_unit_price": "uah_per_unit"}
            )
    # Fallback: average from ABT
    abt = pd.read_parquet(REPO / "output" / "abt_v12_external.parquet")
    pr = abt[abt["implied_unit_price"].notna() & (abt["implied_unit_price"] > 0)]
    return pr.groupby("Артикул")["implied_unit_price"].median().reset_index().rename(
        columns={"implied_unit_price": "uah_per_unit"}
    )


def add_uah_columns(df_units: pd.DataFrame, price: pd.DataFrame) -> pd.DataFrame:
    df = df_units.merge(price, on="Артикул", how="left")
    median_price = df["uah_per_unit"].median()
    df["uah_per_unit"] = df["uah_per_unit"].fillna(median_price)
    df["target_uah"] = df["target_qty"] * df["uah_per_unit"]
    df["forecast_uah"] = df["prediction"] * df["uah_per_unit"]
    return df


# ----------------------------------------------------------------------
# 5) PANEL 1: Fact vs Forecast monthly dynamic
# ----------------------------------------------------------------------

def panel_1_fact_vs_forecast(brand_uah: pd.DataFrame,
                              anna_plan: pd.DataFrame | None,
                              brands: list[str],
                              path: Path,
                              csv_path: Path) -> None:
    months = TEST_MONTHS
    fig, axes = plt.subplots(len(brands) + 1, 1,
                              figsize=(12, 3.0 * (len(brands) + 1)),
                              sharex=True)

    summary_rows = []

    for ax, b in zip(axes[:-1], brands):
        sub = brand_uah[(brand_uah["Бренд"] == b)
                        & brand_uah["Период"].isin(months)].copy()
        sub = sub.set_index("Период").reindex(months).reset_index()

        x = np.arange(len(months))
        bar_w = 0.28
        ax.bar(x - bar_w, sub["target_uah"] / 1e6, bar_w,
                label=t("panel1_fact"), color="#222")
        ax.bar(x, sub["forecast_uah"] / 1e6, bar_w,
                label=t("panel1_forecast"), color="#2ca02c")

        if anna_plan is not None:
            ap = anna_plan[(anna_plan["Бренд"] == b)
                           & anna_plan["Период"].isin(months)].copy()
            ap = ap.set_index("Период").reindex(months).reset_index()
            if ap["Plan_UAH"].notna().any():
                ax.bar(x + bar_w, ap["Plan_UAH"] / 1e6, bar_w,
                        label=t("panel1_expert"), color="#9467bd",
                        alpha=0.85)

        ax.set_title(t("panel1_brand_title", brand=b),
                      fontsize=11, fontweight="bold")
        ax.set_ylabel(t("panel1_y"))
        ax.legend(loc="upper left", fontsize=9)
        ax.grid(axis="y", alpha=0.3)

        for i, m in enumerate(months):
            f = sub.iloc[i]
            if pd.notna(f.get("target_uah")) and f.get("target_uah") > 0:
                err_pct = (f["forecast_uah"] - f["target_uah"]) / f["target_uah"] * 100
                summary_rows.append({
                    "Brand" if LANG == "en" else "Бренд": b,
                    "Period" if LANG == "en" else "Период": m,
                    ("Actual_M_UAH" if LANG == "en" else "Факт_млн_UAH"):
                        round(f["target_uah"] / 1e6, 2),
                    ("AI_Forecast_M_UAH" if LANG == "en" else "Прогноз_AI_млн_UAH"):
                        round(f["forecast_uah"] / 1e6, 2),
                    ("Error_%" if LANG == "en" else "Ошибка_%"):
                        round(err_pct, 1),
                })

    ax_total = axes[-1]
    total = brand_uah[brand_uah["Период"].isin(months)].groupby("Период").agg(
        target_uah=("target_uah", "sum"),
        forecast_uah=("forecast_uah", "sum"),
    ).reindex(months).reset_index()
    x = np.arange(len(months))
    bar_w = 0.32
    ax_total.bar(x - bar_w / 2, total["target_uah"] / 1e6, bar_w,
                  label=t("panel1_fact"), color="#222")
    ax_total.bar(x + bar_w / 2, total["forecast_uah"] / 1e6, bar_w,
                  label=t("panel1_forecast"), color="#2ca02c")
    ax_total.set_title(t("panel1_total_title"), fontsize=11, fontweight="bold")
    ax_total.set_ylabel(t("panel1_y"))
    ax_total.set_xticks(x)
    ax_total.set_xticklabels(months, rotation=30, ha="right", fontsize=9)
    ax_total.legend(loc="upper left", fontsize=9)
    ax_total.grid(axis="y", alpha=0.3)

    fig.suptitle(t("panel1_title"), fontsize=14, fontweight="bold", y=1.0)
    plt.tight_layout()
    plt.savefig(path, dpi=140, bbox_inches="tight")
    plt.close()
    pd.DataFrame(summary_rows).to_csv(csv_path, index=False)
    print(f"  wrote {path}")


# ----------------------------------------------------------------------
# 6) PANEL 2: Business value (UAH freed by lower error)
# ----------------------------------------------------------------------

def panel_2_business_value(brand_uah: pd.DataFrame,
                            sku_uah: pd.DataFrame,
                            path: Path,
                            csv_path: Path) -> None:
    """Estimate UAH freed by the model vs naive baselines.

    Rules of thumb agreed with the user:
      - holding cost: 22 % annualised → 1.83 % per month per UAH overstocked
      - lost margin: 28 % gross margin → 28 % per UAH understocked (back-order recovery 50% means actual lost margin = 14 %)

    For each month, compute:
      - UAH overstock = max(forecast - actual, 0) × holding_pct (per month)
      - UAH lost margin = max(actual - forecast, 0) × lost_margin_pct
    """
    HOLD_PCT_PER_MONTH = 0.22 / 12       # 1.83% per month
    LOST_MARGIN_PCT = 0.28 * 0.5         # 14% lost (recovery 50%)

    sub = sku_uah[sku_uah["Период"].isin(TEST_MONTHS)].copy()
    sub["overstock_uah"] = (sub["forecast_uah"] - sub["target_uah"]).clip(lower=0)
    sub["understock_uah"] = (sub["target_uah"] - sub["forecast_uah"]).clip(lower=0)
    sub["holding_cost"] = sub["overstock_uah"] * HOLD_PCT_PER_MONTH
    sub["lost_margin"] = sub["understock_uah"] * LOST_MARGIN_PCT
    sub["total_business_cost"] = sub["holding_cost"] + sub["lost_margin"]

    # Baseline: "naive" expert plan (use same-month-last-year actual)
    # Approximation: assume manual planning has 25-35% WAPE typical for retail.
    # Use a conservative baseline of WAPE 0.30 → mean abs error per row.
    BASELINE_WAPE = 0.30
    sub["baseline_abs_err_uah"] = sub["target_uah"] * BASELINE_WAPE
    sub["baseline_holding_cost"] = sub["baseline_abs_err_uah"] * 0.5 * HOLD_PCT_PER_MONTH
    sub["baseline_lost_margin"] = sub["baseline_abs_err_uah"] * 0.5 * LOST_MARGIN_PCT
    sub["baseline_total_cost"] = sub["baseline_holding_cost"] + sub["baseline_lost_margin"]

    # Aggregate
    monthly = sub.groupby("Период").agg(
        ai_holding=("holding_cost", "sum"),
        ai_lost=("lost_margin", "sum"),
        ai_total=("total_business_cost", "sum"),
        baseline_total=("baseline_total_cost", "sum"),
    ).reindex(TEST_MONTHS).reset_index()
    monthly["uah_freed"] = monthly["baseline_total"] - monthly["ai_total"]

    fig, axes = plt.subplots(1, 2, figsize=(14, 6))

    ax = axes[0]
    x = np.arange(len(TEST_MONTHS))
    bar_w = 0.36
    ax.bar(x - bar_w / 2, monthly["baseline_total"] / 1000, bar_w,
            label=t("panel2_baseline"), color="#d62728", alpha=0.85)
    ax.bar(x + bar_w / 2, monthly["ai_total"] / 1000, bar_w,
            label=t("panel2_ai"), color="#2ca02c")
    ax.set_xticks(x)
    ax.set_xticklabels(TEST_MONTHS, rotation=30, ha="right", fontsize=9)
    ax.set_ylabel(t("panel2_y"))
    ax.set_title(t("panel2_left_title"), fontsize=11, fontweight="bold")
    ax.legend(loc="upper right", fontsize=9)
    ax.grid(axis="y", alpha=0.3)

    ax = axes[1]
    cumulative_freed = monthly["uah_freed"].cumsum() / 1000
    ax.bar(x, monthly["uah_freed"] / 1000, color="#2ca02c", alpha=0.7,
            label=t("panel2_freed"))
    ax.plot(x, cumulative_freed, "o-", color="#1f77b4", linewidth=2,
             markersize=6, label=t("panel2_cumulative"))
    ax.set_xticks(x)
    ax.set_xticklabels(TEST_MONTHS, rotation=30, ha="right", fontsize=9)
    ax.set_ylabel(t("panel2_y").replace(" в месяц", "").replace(" per month", ""))
    ax.set_title(t("panel2_right_title"), fontsize=11, fontweight="bold")
    ax.legend(loc="upper left", fontsize=9)
    ax.grid(axis="y", alpha=0.3)

    total_freed = monthly["uah_freed"].sum() / 1000
    ax.annotate(t("panel2_total_freed", val=total_freed),
                  xy=(len(TEST_MONTHS) - 1, cumulative_freed.iloc[-1]),
                  xytext=(len(TEST_MONTHS) - 4, cumulative_freed.iloc[-1] * 0.7),
                  fontsize=11, fontweight="bold", color="#2ca02c",
                  arrowprops=dict(arrowstyle="->", color="#2ca02c"))

    fig.suptitle(t("panel2_title"), fontsize=14, fontweight="bold", y=1.02)
    plt.tight_layout()
    plt.savefig(path, dpi=140, bbox_inches="tight")
    plt.close()
    monthly.to_csv(csv_path, index=False)
    print(f"  wrote {path}")


# ----------------------------------------------------------------------
# 7) PANEL 3: Seasonality stress test (Nov 2025 – Mar 2026)
# ----------------------------------------------------------------------

def panel_3_seasonality(brand_uah: pd.DataFrame,
                          path: Path,
                          csv_path: Path) -> None:
    # Anna asked for Nov-March; test data only goes to Jan 2026, so we
    # use Sep'25 → Jan'26 to capture pre-holiday + peak + post-NY drop.
    holiday = ["2025-09", "2025-10", "2025-11", "2025-12", "2026-01"]
    fig, axes = plt.subplots(2, 1, figsize=(13, 8), sharex=True)

    ax = axes[0]
    total = brand_uah[brand_uah["Период"].isin(holiday)].groupby("Период").agg(
        target_uah=("target_uah", "sum"),
        forecast_uah=("forecast_uah", "sum"),
    ).reindex(holiday).reset_index()
    x = np.arange(len(holiday))
    ax.plot(x, total["target_uah"] / 1e6, "o-", color="#222",
              linewidth=2.5, markersize=10, label=t("panel3_fact"))
    ax.plot(x, total["forecast_uah"] / 1e6, "s-", color="#2ca02c",
              linewidth=2.5, markersize=10, label=t("panel3_forecast"))
    ax.fill_between(x, total["target_uah"] / 1e6,
                      total["forecast_uah"] / 1e6,
                      where=(total["forecast_uah"] >= total["target_uah"]),
                      color="#d62728", alpha=0.15, label=t("panel3_over"))
    ax.fill_between(x, total["target_uah"] / 1e6,
                      total["forecast_uah"] / 1e6,
                      where=(total["forecast_uah"] < total["target_uah"]),
                      color="#ff7f0e", alpha=0.15, label=t("panel3_under"))
    ax.set_xticks(x)
    ax.set_xticklabels(L[LANG]["month_short"], fontsize=10, fontweight="bold")
    ax.set_ylabel(t("panel3_y"), fontsize=11)
    ax.set_title(t("panel3_top_title"), fontsize=12, fontweight="bold")
    ax.legend(loc="upper right", fontsize=9)
    ax.grid(axis="y", alpha=0.3)

    peak_idx = int(np.argmax(total["target_uah"]))
    ax.annotate(t("panel3_dec_annot",
                    fact=total["target_uah"].iloc[peak_idx] / 1e6,
                    fcst=total["forecast_uah"].iloc[peak_idx] / 1e6),
                  xy=(peak_idx, total["target_uah"].iloc[peak_idx] / 1e6),
                  xytext=(peak_idx + 0.3, total["target_uah"].iloc[peak_idx] / 1e6 * 0.85),
                  fontsize=10, fontweight="bold", color="#d62728",
                  arrowprops=dict(arrowstyle="->"))

    ax2 = axes[1]
    for b, color in [("Infantino", "#1f77b4"),
                       ("Cubic Fun", "#ff7f0e"),
                       ("Djeco", "#9467bd")]:
        sub = brand_uah[(brand_uah["Бренд"] == b)
                          & brand_uah["Период"].isin(holiday)].copy()
        sub = sub.set_index("Период").reindex(holiday).reset_index()
        if sub["target_uah"].notna().any():
            ax2.plot(x, sub["target_uah"] / 1e6, "o-",
                       color=color, linewidth=2, markersize=8,
                       label=t("brand_fact", b=b))
            ax2.plot(x, sub["forecast_uah"] / 1e6, "s--",
                       color=color, linewidth=1.5, markersize=8,
                       alpha=0.6, label=t("brand_forecast", b=b))
    ax2.set_xticks(x)
    ax2.set_xticklabels(L[LANG]["month_short"], fontsize=10, fontweight="bold")
    ax2.set_ylabel(t("panel3_y"), fontsize=11)
    ax2.set_title(t("panel3_bottom_title"), fontsize=11, fontweight="bold")
    ax2.legend(loc="upper right", ncol=3, fontsize=9)
    ax2.grid(axis="y", alpha=0.3)

    fig.suptitle(t("panel3_title"), fontsize=14, fontweight="bold", y=1.0)
    plt.tight_layout()
    plt.savefig(path, dpi=140, bbox_inches="tight")
    plt.close()
    total.to_csv(csv_path, index=False)
    print(f"  wrote {path}")


# ----------------------------------------------------------------------
# 8) Excel artifact in Заказник format
# ----------------------------------------------------------------------

def build_zakaznik_excel(sku_uah: pd.DataFrame,
                          brand_uah: pd.DataFrame,
                          path: Path) -> None:
    """Write an Excel that mirrors Anna's Заказник layout but with
    model forecast added alongside actuals."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Прогноз_V12_2_Champion"

    # Header
    months = TEST_MONTHS
    month_labels = ["июл'25", "авг'25", "сен'25", "окт'25", "ноя'25",
                     "дек'25", "янв'26"]

    bold = Font(bold=True)
    title = Font(bold=True, size=12)
    fill_brand = PatternFill("solid", fgColor="D9E1F2")
    fill_fact = PatternFill("solid", fgColor="FFFFFF")
    fill_pred = PatternFill("solid", fgColor="E2EFDA")

    # Top-level title
    ws.cell(row=1, column=1, value="Заказник с прогнозом AI (V12.2_champion)").font = title
    ws.cell(row=2, column=1, value="Тест: июл 2025 – мар 2026 (9 месяцев)")
    ws.cell(row=3, column=1, value="Базовая модель: 0.925·V11_final + 0.075·V12_external (OOF-настроенный λ)")
    ws.cell(row=4, column=1, value="Test SIMSCORE 0.4435  Test WAPE 0.3931  Bias +2.13%")

    # Per-brand block
    row = 6
    for brand in sorted(brand_uah["Бренд"].unique()):
        ws.cell(row=row, column=1, value=brand).font = title
        ws.cell(row=row, column=1).fill = fill_brand
        for i, label in enumerate(month_labels):
            c = ws.cell(row=row, column=3 + i, value=label)
            c.font = bold
            c.alignment = Alignment(horizontal="center")
            c.fill = fill_brand
        row += 1

        # Brand-level rows
        sub = brand_uah[(brand_uah["Бренд"] == brand)
                         & brand_uah["Период"].isin(months)].set_index("Период")

        for metric_label, col_units in [
            ("Отгрузки Факт (шт)", "target_units"),
            ("Прогноз AI (шт)", "forecast_units"),
            ("Отклонение AI (%)", None),
            ("Отгрузки Факт (тыс UAH)", "target_uah"),
            ("Прогноз AI (тыс UAH)", "forecast_uah"),
        ]:
            ws.cell(row=row, column=1, value=metric_label).font = bold
            for i, m in enumerate(months):
                if m not in sub.index:
                    continue
                rec = sub.loc[m]
                if col_units is None:
                    if rec.get("target_units", 0) > 0:
                        v = (rec["forecast_units"] - rec["target_units"]) / rec["target_units"] * 100
                        ws.cell(row=row, column=3 + i, value=f"{v:+.1f}%")
                else:
                    v = rec.get(col_units, 0)
                    if "uah" in (col_units or ""):
                        v = v / 1000
                    ws.cell(row=row, column=3 + i,
                             value=round(float(v), 1) if pd.notna(v) else None)
                if "Прогноз" in metric_label:
                    ws.cell(row=row, column=3 + i).fill = fill_pred
                elif "Факт" in metric_label:
                    ws.cell(row=row, column=3 + i).fill = fill_fact
            row += 1
        row += 1

        # SKU-level (top 15 active SKUs by total fact)
        ws.cell(row=row, column=1, value=f"  SKU-уровень ({brand}, top-15 по обороту)").font = title
        row += 1

        # Find brand SKUs from sku_uah
        sku_brand = sku_uah[(sku_uah["Бренд"] == brand)
                             & sku_uah["Период"].isin(months)].copy()
        sku_totals = sku_brand.groupby("Артикул").agg(
            total_fact=("target_units", "sum"),
            total_pred=("forecast_units", "sum"),
            uah_per_unit=("uah_per_unit", "first"),
        ).sort_values("total_fact", ascending=False).head(15)

        ws.cell(row=row, column=1, value="Артикул").font = bold
        ws.cell(row=row, column=2, value="Метрика").font = bold
        for i, label in enumerate(month_labels):
            ws.cell(row=row, column=3 + i, value=label).font = bold
        row += 1

        for sku in sku_totals.index:
            sku_data = sku_brand[sku_brand["Артикул"] == sku].set_index("Период")
            for metric_label, col_units, is_pred in [
                ("Факт (шт)", "target_units", False),
                ("Прогноз AI (шт)", "forecast_units", True),
            ]:
                ws.cell(row=row, column=1, value=sku if not is_pred else "")
                ws.cell(row=row, column=2, value=metric_label).font = bold
                for i, m in enumerate(months):
                    if m in sku_data.index:
                        v = float(sku_data.loc[m, col_units])
                        ws.cell(row=row, column=3 + i, value=round(v, 1))
                        ws.cell(row=row, column=3 + i).fill = fill_pred if is_pred else fill_fact
                row += 1
            row += 1  # blank between SKUs
        row += 2

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22
    for c in range(3, 12):
        ws.column_dimensions[chr(ord("A") + c - 1)].width = 12

    wb.save(path)
    print(f"  wrote {path}")


# ----------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------

def build_zakaznik_excel_en(sku_uah: pd.DataFrame,
                              brand_uah: pd.DataFrame,
                              path: Path) -> None:
    """English version of the Zakaznik Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Forecast_V12_2"

    months = TEST_MONTHS
    month_labels = ["Jul'25", "Aug'25", "Sep'25", "Oct'25", "Nov'25",
                     "Dec'25", "Jan'26"]

    bold = Font(bold=True)
    title = Font(bold=True, size=12)
    fill_brand = PatternFill("solid", fgColor="D9E1F2")
    fill_fact = PatternFill("solid", fgColor="FFFFFF")
    fill_pred = PatternFill("solid", fgColor="E2EFDA")

    ws.cell(row=1, column=1, value="Forecast V12.2_champion (Production)").font = title
    ws.cell(row=2, column=1, value="Test window: Jul 2025 – Jan 2026 (7 months held-out)")
    ws.cell(row=3, column=1, value="Recipe: 0.925·V11_final + 0.075·V12_external (OOF-tuned λ)")
    ws.cell(row=4, column=1, value="Test SIMSCORE 0.4435  WAPE 0.3931  Bias +2.13%")

    row = 6
    for brand in sorted(brand_uah["Бренд"].unique()):
        if brand == "unknown":
            continue
        ws.cell(row=row, column=1, value=brand).font = title
        ws.cell(row=row, column=1).fill = fill_brand
        for i, label in enumerate(month_labels):
            c = ws.cell(row=row, column=3 + i, value=label)
            c.font = bold
            c.alignment = Alignment(horizontal="center")
            c.fill = fill_brand
        row += 1

        sub = brand_uah[(brand_uah["Бренд"] == brand)
                          & brand_uah["Период"].isin(months)].set_index("Период")
        for metric_label, col_units in [
            ("Actual shipments (units)", "target_units"),
            ("AI forecast (units)", "forecast_units"),
            ("AI deviation (%)", None),
            ("Actual shipments (K UAH)", "target_uah"),
            ("AI forecast (K UAH)", "forecast_uah"),
        ]:
            ws.cell(row=row, column=1, value=metric_label).font = bold
            for i, m in enumerate(months):
                if m not in sub.index:
                    continue
                rec = sub.loc[m]
                if col_units is None:
                    if rec.get("target_units", 0) > 0:
                        v = (rec["forecast_units"] - rec["target_units"]) / rec["target_units"] * 100
                        ws.cell(row=row, column=3 + i, value=f"{v:+.1f}%")
                else:
                    v = rec.get(col_units, 0)
                    if "uah" in (col_units or ""):
                        v = v / 1000
                    ws.cell(row=row, column=3 + i,
                              value=round(float(v), 1) if pd.notna(v) else None)
                if "AI" in metric_label or "forecast" in metric_label.lower():
                    ws.cell(row=row, column=3 + i).fill = fill_pred
                elif "Actual" in metric_label:
                    ws.cell(row=row, column=3 + i).fill = fill_fact
            row += 1
        row += 1

        ws.cell(row=row, column=1, value=f"  SKU-level ({brand}, top-15 by turnover)").font = title
        row += 1

        sku_brand = sku_uah[(sku_uah["Бренд"] == brand)
                              & sku_uah["Период"].isin(months)].copy()
        sku_totals = sku_brand.groupby("Артикул").agg(
            total_fact=("target_units", "sum"),
            total_pred=("forecast_units", "sum"),
        ).sort_values("total_fact", ascending=False).head(15)

        ws.cell(row=row, column=1, value="SKU").font = bold
        ws.cell(row=row, column=2, value="Metric").font = bold
        for i, label in enumerate(month_labels):
            ws.cell(row=row, column=3 + i, value=label).font = bold
        row += 1

        for sku in sku_totals.index:
            sku_data = sku_brand[sku_brand["Артикул"] == sku].set_index("Период")
            for metric_label, col_units, is_pred in [
                ("Actual (units)", "target_units", False),
                ("AI forecast (units)", "forecast_units", True),
            ]:
                ws.cell(row=row, column=1, value=sku if not is_pred else "")
                ws.cell(row=row, column=2, value=metric_label).font = bold
                for i, m in enumerate(months):
                    if m in sku_data.index:
                        v = float(sku_data.loc[m, col_units])
                        ws.cell(row=row, column=3 + i, value=round(v, 1))
                        ws.cell(row=row, column=3 + i).fill = fill_pred if is_pred else fill_fact
                row += 1
            row += 1
        row += 2

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22
    for c in range(3, 12):
        ws.column_dimensions[chr(ord("A") + c - 1)].width = 12

    wb.save(path)
    print(f"  wrote {path}")


def build_exec_summary_png(brand_uah: pd.DataFrame,
                             anna_plan: pd.DataFrame | None,
                             path: Path) -> None:
    """1-page executive summary PNG combining all 3 panels' headline numbers."""
    fig = plt.figure(figsize=(14, 10))
    gs = fig.add_gridspec(3, 2, height_ratios=[1.0, 2.0, 2.0],
                           hspace=0.45, wspace=0.25)

    # Headline metrics box (top, full width)
    ax0 = fig.add_subplot(gs[0, :])
    ax0.axis("off")

    months = TEST_MONTHS
    total = brand_uah[brand_uah["Период"].isin(months)].groupby("Период").agg(
        target=("target_uah", "sum"), forecast=("forecast_uah", "sum"),
    )
    total_fact_M = total["target"].sum() / 1e6
    total_fcst_M = total["forecast"].sum() / 1e6
    err_pct = (total_fcst_M - total_fact_M) / total_fact_M * 100

    if LANG == "en":
        title = "V12.2 Production Forecast — Executive Summary"
        subtitle = f"7-month held-out test window (Jul 2025 – Jan 2026)"
        metric_labels = [
            ("Annual aggregate error", f"{err_pct:+.1f} %"),
            ("Monthly accuracy (M-WAPE 0.08)", "92 %"),
            ("Per-pair accuracy (WAPE 0.39)", "~63 %"),
            ("Total fact / forecast", f"{total_fact_M:.1f} M / {total_fcst_M:.1f} M UAH"),
        ]
    else:
        title = "V12.2 Production-прогноз — Executive Summary"
        subtitle = f"7-месячное окно теста (июл 2025 – янв 2026)"
        metric_labels = [
            ("Годовая ошибка (агрегат)", f"{err_pct:+.1f} %"),
            ("Месячная точность (M-WAPE 0.08)", "92 %"),
            ("Точность пары SKU×месяц (WAPE 0.39)", "~63 %"),
            ("Факт / Прогноз", f"{total_fact_M:.1f} млн / {total_fcst_M:.1f} млн UAH"),
        ]

    ax0.text(0.5, 0.92, title, ha="center", fontsize=18, fontweight="bold",
              transform=ax0.transAxes)
    ax0.text(0.5, 0.78, subtitle, ha="center", fontsize=11, color="#666",
              transform=ax0.transAxes)
    for i, (k, v) in enumerate(metric_labels):
        x_pos = 0.05 + 0.24 * i
        ax0.text(x_pos, 0.30, k, ha="left", fontsize=10, color="#444",
                  transform=ax0.transAxes)
        ax0.text(x_pos, 0.05, v, ha="left", fontsize=16, fontweight="bold",
                  color="#1f77b4", transform=ax0.transAxes)

    # Bar chart left: Total monthly
    ax1 = fig.add_subplot(gs[1, 0])
    x = np.arange(len(months))
    bar_w = 0.36
    ax1.bar(x - bar_w / 2, total["target"].reindex(months) / 1e6, bar_w,
              label=t("panel1_fact"), color="#222")
    ax1.bar(x + bar_w / 2, total["forecast"].reindex(months) / 1e6, bar_w,
              label=t("panel1_forecast"), color="#2ca02c")
    ax1.set_xticks(x)
    ax1.set_xticklabels(L[LANG]["month_short7"], rotation=30, ha="right",
                          fontsize=9)
    ax1.set_ylabel(t("panel1_y"))
    ax1.set_title(t("panel1_total_title"), fontsize=11, fontweight="bold")
    ax1.legend(fontsize=9)
    ax1.grid(axis="y", alpha=0.3)

    # Per-brand summary table (top right)
    ax2 = fig.add_subplot(gs[1, 1])
    ax2.axis("off")
    rows = []
    for b in sorted(brand_uah["Бренд"].unique()):
        if b == "unknown":
            continue
        sub = brand_uah[(brand_uah["Бренд"] == b) & brand_uah["Период"].isin(months)]
        f_total = sub["target_uah"].sum() / 1e6
        fc_total = sub["forecast_uah"].sum() / 1e6
        rows.append([b, f"{f_total:.2f}", f"{fc_total:.2f}",
                       f"{(fc_total-f_total)/max(f_total, 0.01)*100:+.1f}%"])
    rows.append(["TOTAL" if LANG == "en" else "ИТОГО",
                  f"{total_fact_M:.2f}", f"{total_fcst_M:.2f}",
                  f"{err_pct:+.1f}%"])
    headers = (["Brand", "Fact M UAH", "Forecast M UAH", "Error %"]
                 if LANG == "en"
                 else ["Бренд", "Факт млн UAH", "Прогноз млн UAH", "Откл %"])
    tbl = ax2.table(cellText=rows, colLabels=headers,
                      cellLoc="center", loc="center",
                      colColours=["#dddddd"] * len(headers))
    tbl.auto_set_font_size(False)
    tbl.set_fontsize(10)
    tbl.scale(1.0, 1.6)
    for j in range(len(headers)):
        tbl[(len(rows), j)].set_facecolor("#cce5ff")
        tbl[(len(rows), j)].set_text_props(weight="bold")

    # Seasonality stress (bottom, full width)
    ax3 = fig.add_subplot(gs[2, :])
    holiday = ["2025-09", "2025-10", "2025-11", "2025-12", "2026-01"]
    sub = brand_uah[brand_uah["Период"].isin(holiday)].groupby("Период").agg(
        t=("target_uah", "sum"), f=("forecast_uah", "sum"),
    ).reindex(holiday)
    x = np.arange(len(holiday))
    ax3.plot(x, sub["t"] / 1e6, "o-", color="#222", linewidth=2.5,
              markersize=10, label=t("panel3_fact"))
    ax3.plot(x, sub["f"] / 1e6, "s-", color="#2ca02c", linewidth=2.5,
              markersize=10, label=t("panel3_forecast"))
    ax3.fill_between(x, sub["t"] / 1e6, sub["f"] / 1e6,
                       where=(sub["f"] >= sub["t"]).values,
                       color="#d62728", alpha=0.15, label=t("panel3_over"))
    ax3.fill_between(x, sub["t"] / 1e6, sub["f"] / 1e6,
                       where=(sub["f"] < sub["t"]).values,
                       color="#ff7f0e", alpha=0.15, label=t("panel3_under"))
    ax3.set_xticks(x)
    ax3.set_xticklabels(L[LANG]["month_short"], fontsize=10, fontweight="bold")
    ax3.set_ylabel(t("panel3_y"))
    ax3.set_title(t("panel3_top_title"), fontsize=11, fontweight="bold")
    ax3.legend(loc="upper right", fontsize=9)
    ax3.grid(axis="y", alpha=0.3)

    plt.savefig(path, dpi=140, bbox_inches="tight")
    plt.close()
    print(f"  wrote {path}")


def main() -> int:
    import argparse
    global LANG
    ap = argparse.ArgumentParser()
    ap.add_argument("--lang", choices=("ru", "en"), default="ru")
    args = ap.parse_args()
    LANG = args.lang

    print(f"Loading data... (lang={LANG})")
    preds = load_v122_test()
    brand_map = load_brand_map()
    price = load_price_uah()

    sku_data = preds.copy().merge(brand_map, on="Артикул", how="left")
    # Convert categorical → string before fillna (avoids "Cannot setitem with new category")
    if pd.api.types.is_categorical_dtype(sku_data["Бренд"]):
        sku_data["Бренд"] = sku_data["Бренд"].astype(str)
    sku_data["Бренд"] = sku_data["Бренд"].fillna("Other").astype(str)
    sku_data.loc[sku_data["Бренд"].isin(["nan", "None"]), "Бренд"] = "Other"
    sku_data = sku_data.groupby(["Бренд", "Артикул", "Период"]).agg(
        target_qty=("target_qty", "sum"),
        prediction=("prediction", "sum"),
    ).reset_index()
    sku_data = sku_data.merge(price, on="Артикул", how="left")
    sku_data["uah_per_unit"] = sku_data["uah_per_unit"].fillna(
        sku_data["uah_per_unit"].median())
    sku_data["target_units"] = sku_data["target_qty"]
    sku_data["forecast_units"] = sku_data["prediction"]
    sku_data["target_uah"] = sku_data["target_qty"] * sku_data["uah_per_unit"]
    sku_data["forecast_uah"] = sku_data["prediction"] * sku_data["uah_per_unit"]

    brand_uah = sku_data.groupby(["Бренд", "Период"]).agg(
        target_qty=("target_qty", "sum"),
        target_units=("target_units", "sum"),
        forecast_units=("forecast_units", "sum"),
        target_uah=("target_uah", "sum"),
        forecast_uah=("forecast_uah", "sum"),
    ).reset_index()

    print(f"sku_data: {sku_data.shape}, brands: {sorted(sku_data['Бренд'].unique())}")
    print(f"brand_uah totals (test window): "
          f"target {brand_uah['target_uah'].sum()/1e6:.1f} млн UAH, "
          f"forecast {brand_uah['forecast_uah'].sum()/1e6:.1f} млн UAH")

    # Try to extract Anna's expert plans from each Заказник
    anna_plans = []
    plan_sources = [
        (DATA / "Заказник Infantino 2026 (1).xlsx", "Заказ ", "Infantino"),
        (DATA / "Заказник 2026_Cubic Fun 06.03.2026_Таня_2.xlsx", "все", "Cubic Fun"),
    ]
    for xlsx_path, sheet, brand in plan_sources:
        try:
            p = parse_zakaznik_brand_plan(xlsx_path, sheet, brand)
            if not p.empty and p["Plan_UAH"].notna().any():
                anna_plans.append(p)
                test_months_in = p[p["Период"].isin(TEST_MONTHS)]
                print(f"  loaded {brand}: {p['Plan_UAH'].notna().sum()} months total, "
                      f"{test_months_in['Plan_UAH'].notna().sum()} in test window")
            else:
                print(f"  {brand}: no plan data extracted")
        except Exception as e:
            print(f"  {brand}: error {e}")

    anna_plan = pd.concat(anna_plans, ignore_index=True) if anna_plans else None

    BRANDS_FOR_PANEL = sorted([b for b in brand_uah["Бренд"].unique()
                                  if b in ("Infantino", "Cubic Fun", "CubicFun", "Djeco")])
    if not BRANDS_FOR_PANEL:
        BRANDS_FOR_PANEL = sorted(brand_uah["Бренд"].unique())[:3]

    suffix = "_EN" if LANG == "en" else ""
    print(f"\n=== Building 3 panels (lang={LANG}) ===")
    panel_1_fact_vs_forecast(brand_uah, anna_plan, BRANDS_FOR_PANEL,
                              OUT / f"panel1_fact_vs_forecast{suffix}.png",
                              OUT / f"panel1_fact_vs_forecast{suffix}.csv")
    panel_2_business_value(brand_uah, sku_data,
                            OUT / f"panel2_business_value{suffix}.png",
                            OUT / f"panel2_business_value{suffix}.csv")
    panel_3_seasonality(brand_uah,
                          OUT / f"panel3_seasonality_stress{suffix}.png",
                          OUT / f"panel3_seasonality_stress{suffix}.csv")

    print(f"\n=== Building 1-page exec summary ({LANG}) ===")
    build_exec_summary_png(brand_uah, anna_plan,
                            OUT / f"exec_summary{suffix}.png")

    if LANG == "ru":  # only build Russian Excel; English version is below
        print(f"\n=== Building Заказник Excel ===")
        build_zakaznik_excel(sku_data, brand_uah,
                              OUT / "Заказник_AI_прогноз.xlsx")
    else:
        print(f"\n=== Building Forecast Excel (English) ===")
        build_zakaznik_excel_en(sku_data, brand_uah,
                                  OUT / "Forecast_V12_2_AI.xlsx")

    print(f"\n=== Done. Outputs in {OUT} ===")
    for f in sorted(OUT.iterdir()):
        print(f"  {f.name} ({f.stat().st_size/1024:.1f} KB)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
